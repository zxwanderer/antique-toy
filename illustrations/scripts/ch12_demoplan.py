#!/usr/bin/env python3
"""Render n1k-o's GABBA demoplan spreadsheet as a PNG illustration.

Reads _in/demoplan_v5.xlsx and produces a colour-coded visualization
showing the first ~40 rows of Pattern 4 (the most visually interesting
section with all layers active).

Source: n1k-o (4D+TBK), GABBA demo sync workflow, shared 2026-03-01.
"""

import openpyxl
from openpyxl.utils import get_column_letter
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from matplotlib.colors import to_rgba
import numpy as np

# ── Load workbook ──────────────────────────────────────────────
wb = openpyxl.load_workbook('_in/demoplan_v5.xlsx')
ws = wb.active

# ── Configuration ──────────────────────────────────────────────
# Show rows from Pattern4 start (interesting section with all layers)
# From the screenshot: Pattern4 starts around row where frame=1536
START_ROW = None
NUM_ROWS = 40
COLS = list(range(1, 11))  # Pattern through Слой6

# Find Pattern4 start
for row in range(2, ws.max_row + 1):
    val = ws.cell(row, 1).value
    if val and 'Pattern4' in str(val):
        START_ROW = row
        break

if START_ROW is None:
    # Fallback: find frame 1536
    for row in range(2, ws.max_row + 1):
        frame = ws.cell(row, 3).value
        if frame is not None and int(frame) >= 1536:
            START_ROW = row
            break

if START_ROW is None:
    START_ROW = 2

print(f"Rendering from row {START_ROW} (frame={ws.cell(START_ROW, 3).value})")

# ── Extract cell colors (resolve theme to RGB) ────────────────
# Theme color map from the actual xlsx (extracted from screenshot reference)
# openpyxl theme indices → approximate RGB
THEME_COLORS = {
    0: '#FFFFFF',   # white
    1: '#000000',   # black
    2: '#1F497D',   # dark blue
    3: '#EEECE1',   # tan
    4: '#4F81BD',   # blue
    5: '#C0504D',   # red
    6: '#9BBB59',   # olive green
    7: '#8064A2',   # purple
    8: '#4BACC6',   # teal
    9: '#F79646',   # orange
}

def get_cell_color(cell):
    """Return hex color string or None for a cell's background fill."""
    fill = cell.fill
    if fill.patternType != 'solid':
        return None
    fg = fill.fgColor
    if fg is None:
        return None
    # Direct RGB
    if fg.rgb and fg.rgb not in ('00000000', '0') and not str(fg.rgb).startswith('Value'):
        rgb = str(fg.rgb)
        if len(rgb) == 8:  # AARRGGBB
            return '#' + rgb[2:]
        return '#' + rgb
    # Theme-based (we'll try to resolve)
    try:
        theme = fg.theme
        if theme is not None and isinstance(theme, int):
            return THEME_COLORS.get(theme, None)
    except:
        pass
    return None

# ── Collect data ───────────────────────────────────────────────
headers = [str(ws.cell(1, c).value or '') for c in COLS]
data = []
colors = []

for row in range(START_ROW, min(START_ROW + NUM_ROWS, ws.max_row + 1)):
    row_data = []
    row_colors = []
    for col in COLS:
        cell = ws.cell(row, col)
        val = str(cell.value or '')
        if len(val) > 25:
            val = val[:25] + '...'
        row_data.append(val)
        row_colors.append(get_cell_color(cell))
    data.append(row_data)
    colors.append(row_colors)

# ── Render ─────────────────────────────────────────────────────
n_rows = len(data)
n_cols = len(COLS)

# Column widths (proportional)
col_widths = [1.2, 0.7, 0.7, 4.0, 0.8, 0.8, 0.8, 0.8, 0.8, 0.8]
total_w = sum(col_widths)

fig_w = 14
fig_h = max(8, n_rows * 0.22 + 1.5)
fig, ax = plt.subplots(1, 1, figsize=(fig_w, fig_h))
ax.set_xlim(0, total_w)
ax.set_ylim(0, n_rows + 1.5)
ax.invert_yaxis()
ax.axis('off')

cell_h = 1.0

# Header row
x = 0
for j, (hdr, w) in enumerate(zip(headers, col_widths)):
    rect = patches.FancyBboxPatch((x, 0), w, cell_h,
        boxstyle="square,pad=0", facecolor='#D9E1F2', edgecolor='#8DB4E2', linewidth=0.5)
    ax.add_patch(rect)
    ax.text(x + w/2, cell_h/2, hdr, ha='center', va='center',
            fontsize=7, fontweight='bold', fontfamily='monospace')
    x += w

# Data rows
layer_colors = {
    'Ударные': '#1F3864',   # dark navy (drums)
    'Хеты':    '#FFC000',   # gold (hi-hats)
    'Мелодия': '#00B050',   # green (melody)
    'Бас':     '#7030A0',   # purple (bass)
    'Эффект':  '#1F3864',   # dark navy (effects)
    'Слой6':   '#7030A0',   # purple (layer 6)
}

for i, (row_data, row_colors) in enumerate(zip(data, colors)):
    y = (i + 1) * cell_h
    x = 0
    for j, (val, w) in enumerate(zip(row_data, col_widths)):
        # Determine fill color
        fc = '#FFFFFF'
        tc = '#000000'

        # Check if this is a colored layer cell
        col_name = headers[j]
        cell_color = row_colors[j]

        if cell_color:
            fc = cell_color
            # Light text on dark backgrounds
            r, g, b, _ = to_rgba(fc)
            if (r * 0.299 + g * 0.587 + b * 0.114) < 0.5:
                tc = '#FFFFFF'

        # Pattern column highlight (blue rows for pattern headers)
        if j == 0 and val and val.startswith('[Pattern'):
            fc = '#4472C4'
            tc = '#FFFFFF'

        # Alternating row tint
        if fc == '#FFFFFF' and i % 2 == 1:
            fc = '#F8F8F8'

        rect = patches.FancyBboxPatch((x, y), w, cell_h,
            boxstyle="square,pad=0", facecolor=fc, edgecolor='#D0D0D0', linewidth=0.3)
        ax.add_patch(rect)

        fontsize = 5.5 if j == 3 else 6.5  # smaller for Vortex data column
        ax.text(x + 0.05, y + cell_h/2, val, ha='left', va='center',
                fontsize=fontsize, color=tc, fontfamily='monospace')
        x += w

# Title
ax.text(total_w / 2, -0.3,
        "n1k-o's demoplan for Unspoken (4D+TBK) — colour-coded sync map (Pattern 4, frames 1536–1614)",
        ha='center', va='center', fontsize=10, fontweight='bold', style='italic')

# Legend
legend_y = n_rows + 1.2
legend_items = [
    ('Ударные (Drums)', '#1F3864'),
    ('Хеты (Hi-hats)', '#FFC000'),
    ('Мелодия (Melody)', '#00B050'),
    ('Бас (Bass)', '#7030A0'),
    ('Эффект (FX)', '#1F3864'),
]
x_legend = 1.0
for label, color in legend_items:
    rect = patches.FancyBboxPatch((x_legend, legend_y), 0.5, 0.5,
        boxstyle="square,pad=0", facecolor=color, edgecolor='#666666', linewidth=0.5)
    ax.add_patch(rect)
    ax.text(x_legend + 0.6, legend_y + 0.25, label, ha='left', va='center', fontsize=7)
    x_legend += 2.2

plt.tight_layout()
outpath = 'illustrations/output/ch12_demoplan.png'
fig.savefig(outpath, dpi=200, bbox_inches='tight', facecolor='white')
print(f'Saved: {outpath}')
plt.close()
